import os

import openpyxl
import pandas as pd
from bson import ObjectId

from app.enums.parsers import ParserStatus, ParserType
from app.repositories.user_parsers import UserParsersRepository
from app.schemas.parser import BaseParsersSchema
from app.services import parser


async def get_headers(keys, header_row: int, sheet):
    for col, key in enumerate(keys, start=1):
        sheet.cell(row=header_row, column=col, value=key)


async def get_keys(data):
    keys = set()

    for d in data:
        key_set = d.keys()
        keys = keys.union(key_set)
    return keys


async def create_excel(data: list, name_of_excel: str) -> str:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    keys = await get_keys(data)
    await get_headers(keys, 1, sheet)
    for row, obj in enumerate(data, start=1 + 1):
        for col, key in enumerate(keys, start=1):
            value = str(obj.get(key))
            sheet.cell(row=row, column=col, value=value)  # type: ignore
    excel_file = f"{name_of_excel}.xlsx"
    workbook.save(excel_file)
    return excel_file


async def read_headers(excel_file):
    data_frame = pd.read_excel(excel_file)
    headers = data_frame.columns.tolist()
    return headers, data_frame


async def read_excel(excel_headers, df):
    row = 0
    first_row_values = {}
    result = []
    while row is not None:
        try:
            for header in excel_headers:
                first_row_values[header] = df[header].iloc[row]
            row += 1
            result.append(first_row_values)
        except IndexError:
            break
    return result


async def create_excel_file(base_id: str):
    base = await UserParsersRepository().get_by_id(_id=ObjectId(base_id))
    data = base.get("parser_data")
    if base.get("parser_type") == ParserType.avito.value:
        file_name = ParserType.avito.value
        excel_file = await create_excel(data, file_name)  # type: ignore
        return excel_file, file_name
    elif base.get("parser_type") == ParserType.avito.yandex:
        file_name = ParserType.yandex.value
        excel_file = await create_excel(data, file_name)  # type: ignore
        return excel_file, file_name
    elif base.get("parser_type") == ParserType.vk_groups.value:
        file_name = ParserType.vk_groups.value
        excel_file = await create_excel(data, file_name)  # type: ignore
        return excel_file, file_name
    elif base.get("parser_type") == ParserType.vk_posts.value:
        file_name = ParserType.vk_posts.value
        excel_file = await create_excel(data, file_name)  # type: ignore
        return excel_file, file_name


async def delete_file(file_path: str):
    os.remove(file_path)


async def import_excel(owner_id: str, excel_file, parser_type: str, filters: list[str]):
    headers, data_frame = await read_headers(excel_file)
    excel_data = await read_excel(headers, data_frame)
    parser_data = BaseParsersSchema(
        parser_type=parser_type,
        owner_id=owner_id,
        status=ParserStatus.parsed,
        parser_data=excel_data,
        filters=filters
    )
    user_parser = await parser.create_base(parser_data)
    return user_parser
